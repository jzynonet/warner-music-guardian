import csv
import io
from typing import List, Dict
from openpyxl import load_workbook

class BulkImporter:
    """Utility for bulk importing keywords and artists"""
    
    @staticmethod
    def parse_csv(file_content: str) -> List[Dict]:
        """
        Parse CSV file for keyword import
        Expected format: keyword,artist_name,auto_flag,priority
        """
        keywords = []
        reader = csv.DictReader(io.StringIO(file_content))
        
        for row in reader:
            keyword_data = {
                'keyword': row.get('keyword', '').strip(),
                'artist_name': row.get('artist_name', '').strip(),
                'auto_flag': row.get('auto_flag', 'false').lower() in ['true', '1', 'yes'],
                'priority': row.get('priority', 'Medium').strip()
            }
            
            if keyword_data['keyword']:  # Only add if keyword is not empty
                keywords.append(keyword_data)
        
        return keywords
    
    @staticmethod
    def parse_excel(file_path: str) -> List[Dict]:
        """
        Parse Excel file for keyword import
        Expected columns: keyword, artist_name, auto_flag, priority
        """
        keywords = []
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            keyword_data = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    keyword_data[header] = row[idx]
            
            if keyword_data.get('keyword'):
                keywords.append({
                    'keyword': str(keyword_data.get('keyword', '')).strip(),
                    'artist_name': str(keyword_data.get('artist_name', '')).strip(),
                    'auto_flag': str(keyword_data.get('auto_flag', 'false')).lower() in ['true', '1', 'yes'],
                    'priority': str(keyword_data.get('priority', 'Medium')).strip()
                })
        
        return keywords
    
    @staticmethod
    def generate_template_csv() -> str:
        """Generate CSV template for keyword import"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['keyword', 'artist_name', 'auto_flag', 'priority'])
        writer.writerow(['Example Brand Name', 'Artist Name', 'true', 'High'])
        writer.writerow(['Example Product', 'Artist Name', 'false', 'Medium'])
        
        return output.getvalue()
    
    @staticmethod
    def parse_artists_csv(file_content: str) -> List[Dict]:
        """
        Parse CSV file for artist import
        Expected format: name,email,contact_person,notes
        """
        artists = []
        reader = csv.DictReader(io.StringIO(file_content))
        
        for row in reader:
            artist_data = {
                'name': row.get('name', '').strip(),
                'email': row.get('email', '').strip(),
                'contact_person': row.get('contact_person', '').strip(),
                'notes': row.get('notes', '').strip()
            }
            
            if artist_data['name']:  # Only add if name is not empty
                artists.append(artist_data)
        
        return artists
